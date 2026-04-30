"""Tests for Excel export service."""
import io
import pytest
from openpyxl import load_workbook
from app.models import Assessment, User, Response, ToolInventory
from app.services.excel_service import build_customer_excel, build_consultant_excel


@pytest.fixture
def full_assessment(db):
    """Assessment with responses and tools for Excel testing."""
    a = Assessment(
        customer_org="Excel Test Org",
        framework="dod_zt",
        variant="zt_only",
        status="finalized",
    )
    db.session.add(a)
    db.session.flush()

    tool = ToolInventory(assessment_id=a.id, name="Splunk", vendor="Splunk Inc", category="SIEM")
    db.session.add(tool)

    responses = [
        Response(assessment_id=a.id, pillar="user", activity_id="dod_zt.user.1.1",
                 current_state_value="partial", target_state_value="target"),
        Response(assessment_id=a.id, pillar="user", activity_id="dod_zt.user.1.2",
                 current_state_value="not_met", target_state_value="advanced",
                 evidence_notes="No phishing-resistant MFA yet"),
        Response(assessment_id=a.id, pillar="device", activity_id="dod_zt.device.2.1",
                 current_state_value="target", target_state_value="target"),
    ]
    for r in responses:
        db.session.add(r)

    db.session.commit()
    return a


def test_customer_excel_returns_bytes(app, full_assessment):
    with app.app_context():
        data = build_customer_excel(full_assessment)
    assert isinstance(data, bytes)
    assert len(data) > 0


def test_customer_excel_has_expected_sheets(app, full_assessment):
    with app.app_context():
        data = build_customer_excel(full_assessment)
    wb = load_workbook(io.BytesIO(data))
    assert "Executive Summary" in wb.sheetnames
    assert "Gap Register" in wb.sheetnames
    assert "Tool Inventory Mapping" in wb.sheetnames
    assert "Methodology" in wb.sheetnames


def test_customer_excel_has_pillar_sheets(app, full_assessment):
    with app.app_context():
        data = build_customer_excel(full_assessment)
    wb = load_workbook(io.BytesIO(data))
    # DoD ZT has User, Device, Network_Environment, etc. pillars (/ replaced with _)
    assert "User" in wb.sheetnames
    assert "Device" in wb.sheetnames
    assert "Network_Environment" in wb.sheetnames


def test_gap_register_contains_gaps(app, full_assessment):
    with app.app_context():
        data = build_customer_excel(full_assessment)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Gap Register"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    # dod_zt.user.1.1 (partial->target) and dod_zt.user.1.2 (not_met->advanced) are gaps
    activity_ids = [r[1] for r in rows if r[1]]
    assert "dod_zt.user.1.1" in activity_ids
    assert "dod_zt.user.1.2" in activity_ids
    # dod_zt.device.2.1 is target->target, no gap
    assert "dod_zt.device.2.1" not in activity_ids


def test_tool_inventory_sheet(app, full_assessment):
    with app.app_context():
        data = build_customer_excel(full_assessment)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Tool Inventory Mapping"]
    all_text = " ".join(
        str(cell.value) for row in ws.iter_rows() for cell in row if cell.value
    )
    assert "Splunk" in all_text


def test_consultant_excel_has_extra_sheets(app, full_assessment):
    with app.app_context():
        data = build_consultant_excel(full_assessment)
    wb = load_workbook(io.BytesIO(data))
    assert "Admin Notes" in wb.sheetnames
    assert "AI Call Log" in wb.sheetnames
    assert "Audit Log" in wb.sheetnames
    # Should also have customer sheets
    assert "Executive Summary" in wb.sheetnames
    assert "Gap Register" in wb.sheetnames
