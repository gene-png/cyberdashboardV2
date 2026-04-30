"""Unit tests for the ATT&CK coverage Excel generator."""
import io
import pytest
from datetime import datetime, timezone
from unittest.mock import MagicMock

from app.services.attack_coverage_excel import (
    compute_coverage_matrix,
    build_attack_coverage_excel,
)
from app.services.attack_mapper import classify_gap_status


def _make_technique(full_id, tactic="Initial Access", is_sub=False):
    t = MagicMock()
    t.full_id = full_id
    t.technique_id = full_id.split(".")[0]
    t.sub_technique_id = full_id if is_sub else None
    t.name = f"Technique {full_id}"
    t.tactic = tactic
    t.description = f"Description for {full_id}"
    t.url = f"https://attack.mitre.org/techniques/{full_id.replace('.', '/')}"
    t.is_sub_technique = is_sub
    return t


def _make_tool(name, vendor="ACME", category="EDR"):
    t = MagicMock()
    t.name = name
    t.vendor = vendor
    t.category = category
    t.notes = f"{name} notes"
    return t


TECHNIQUES = [
    _make_technique("T1078", "Initial Access"),
    _make_technique("T1098", "Persistence"),
    _make_technique("T1055", "Defense Evasion"),
    _make_technique("T1055.001", "Defense Evasion", is_sub=True),
    _make_technique("T1190", "Initial Access"),
]

COVERAGE_DATA = [
    {
        "tool": _make_tool("ToolA"),
        "activity_ids": ["act.1"],
        "results": [
            {"technique_id": "T1078", "coverage_type": "detect", "confidence": "high", "rationale": "Monitors auth."},
            {"technique_id": "T1098", "coverage_type": "prevent", "confidence": "medium", "rationale": "Blocks cred changes."},
        ],
    },
    {
        "tool": _make_tool("ToolB", vendor="Vendor B"),
        "activity_ids": ["act.2"],
        "results": [
            {"technique_id": "T1078", "coverage_type": "prevent", "confidence": "high", "rationale": "MFA enforcement."},
            {"technique_id": "T1055", "coverage_type": "detect", "confidence": "low", "rationale": "Process injection detection."},
        ],
    },
]

NOW = datetime(2026, 4, 29, 12, 0, 0, tzinfo=timezone.utc)


# ---- compute_coverage_matrix ----

def test_matrix_has_all_techniques():
    matrix = compute_coverage_matrix(COVERAGE_DATA, TECHNIQUES)
    assert len(matrix) == len(TECHNIQUES)
    for t in TECHNIQUES:
        assert t.full_id in matrix

def test_matrix_detect_tools_populated():
    matrix = compute_coverage_matrix(COVERAGE_DATA, TECHNIQUES)
    assert "ToolA" in matrix["T1078"]["detect_tools"]
    assert "ToolB" not in matrix["T1078"]["detect_tools"]

def test_matrix_prevent_tools_populated():
    matrix = compute_coverage_matrix(COVERAGE_DATA, TECHNIQUES)
    assert "ToolB" in matrix["T1078"]["prevent_tools"]

def test_matrix_gap_status_full_when_detect_and_prevent():
    matrix = compute_coverage_matrix(COVERAGE_DATA, TECHNIQUES)
    assert matrix["T1078"]["gap_status"] == "Full"

def test_matrix_gap_status_single_tool_for_t1098():
    matrix = compute_coverage_matrix(COVERAGE_DATA, TECHNIQUES)
    # T1098 has only ToolA with prevent — single tool
    assert matrix["T1098"]["gap_status"] == "Single Tool"

def test_matrix_gap_status_none_for_uncovered():
    matrix = compute_coverage_matrix(COVERAGE_DATA, TECHNIQUES)
    assert matrix["T1190"]["gap_status"] == "None"

def test_matrix_rationale_combined():
    matrix = compute_coverage_matrix(COVERAGE_DATA, TECHNIQUES)
    rationale = matrix["T1078"]["rationale_summary"]
    assert "ToolA" in rationale or "ToolB" in rationale


# ---- build_attack_coverage_excel ----

def test_excel_builds_without_error():
    xlsx = build_attack_coverage_excel(COVERAGE_DATA, TECHNIQUES, NOW, "claude-sonnet-4-6")
    assert isinstance(xlsx, bytes)
    assert len(xlsx) > 0

def test_excel_has_five_sheets():
    from openpyxl import load_workbook
    xlsx = build_attack_coverage_excel(COVERAGE_DATA, TECHNIQUES, NOW, "claude-sonnet-4-6")
    wb = load_workbook(io.BytesIO(xlsx))
    assert len(wb.sheetnames) == 5

def test_excel_sheet_names():
    from openpyxl import load_workbook
    xlsx = build_attack_coverage_excel(COVERAGE_DATA, TECHNIQUES, NOW, "claude-sonnet-4-6")
    wb = load_workbook(io.BytesIO(xlsx))
    names = wb.sheetnames
    assert "Summary" in names
    assert "Coverage Matrix" in names
    assert "Gaps" in names
    assert "Tool Coverage" in names
    assert "Methodology" in names

def test_excel_coverage_matrix_has_all_techniques():
    from openpyxl import load_workbook
    xlsx = build_attack_coverage_excel(COVERAGE_DATA, TECHNIQUES, NOW, "claude-sonnet-4-6")
    wb = load_workbook(io.BytesIO(xlsx))
    ws = wb["Coverage Matrix"]
    technique_ids = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    for t in TECHNIQUES:
        assert t.technique_id in technique_ids

def test_excel_gaps_sheet_excludes_full_coverage():
    from openpyxl import load_workbook
    xlsx = build_attack_coverage_excel(COVERAGE_DATA, TECHNIQUES, NOW, "claude-sonnet-4-6")
    wb = load_workbook(io.BytesIO(xlsx))
    ws = wb["Gaps"]
    gap_statuses = [ws.cell(row=r, column=5).value for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=5).value]
    assert "Full" not in gap_statuses

def test_excel_tool_coverage_has_both_tools():
    from openpyxl import load_workbook
    xlsx = build_attack_coverage_excel(COVERAGE_DATA, TECHNIQUES, NOW, "claude-sonnet-4-6")
    wb = load_workbook(io.BytesIO(xlsx))
    ws = wb["Tool Coverage"]
    tool_names = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value]
    assert "ToolA" in tool_names
    assert "ToolB" in tool_names

def test_excel_methodology_contains_model_name():
    from openpyxl import load_workbook
    xlsx = build_attack_coverage_excel(COVERAGE_DATA, TECHNIQUES, NOW, "claude-sonnet-4-6")
    wb = load_workbook(io.BytesIO(xlsx))
    ws = wb["Methodology"]
    all_text = " ".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value)
    assert "claude-sonnet-4-6" in all_text

def test_excel_summary_contains_generated_date():
    from openpyxl import load_workbook
    xlsx = build_attack_coverage_excel(COVERAGE_DATA, TECHNIQUES, NOW, "claude-sonnet-4-6")
    wb = load_workbook(io.BytesIO(xlsx))
    ws = wb["Summary"]
    all_text = " ".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value)
    assert "2026-04-29" in all_text

def test_excel_with_empty_coverage_data():
    # Should still build without error
    xlsx = build_attack_coverage_excel([], TECHNIQUES, NOW, "claude-sonnet-4-6")
    assert isinstance(xlsx, bytes)

def test_excel_with_no_techniques():
    xlsx = build_attack_coverage_excel(COVERAGE_DATA, [], NOW, "claude-sonnet-4-6")
    assert isinstance(xlsx, bytes)
