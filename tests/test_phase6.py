"""Tests for Phase 6 features: tool mapping workflow, customer terms, workspace resume, Excel."""
import json
import pytest
from datetime import datetime, timezone
from unittest.mock import patch, MagicMock

from tests.conftest import login
from app.models import Assessment, ToolInventory, SensitiveTerm
from app.models.tool_activity_mapping import ToolActivityMapping
from app.models.mapping_suggestions_log import MappingSuggestionsLog
from app.models.mapping_change import MappingChange
from app.extensions import db as ext_db


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _unlock_admin(client):
    with client.session_transaction() as sess:
        sess["admin_unlocked_at"] = datetime.now(timezone.utc).isoformat()


def _make_tool(db, assessment_id, name="Test Tool", vendor="ACME", category="EDR"):
    tool = ToolInventory(
        assessment_id=assessment_id,
        name=name,
        vendor=vendor,
        category=category,
        notes="Test deployment",
    )
    db.session.add(tool)
    db.session.commit()
    return tool


# ---------------------------------------------------------------------------
# Workspace resume (#4)
# ---------------------------------------------------------------------------

def test_workspace_redirects_to_current_step(client, db, sample_assessment):
    sample_assessment.current_step = "pillar_user"
    db.session.commit()

    login(client, "testcustomer", "custpass")
    resp = client.get(
        f"/assessments/{sample_assessment.id}",
        follow_redirects=False,
    )
    assert resp.status_code == 302
    assert "pillar/user" in resp.headers["Location"]


def test_workspace_shows_overview_with_param(client, db, sample_assessment):
    sample_assessment.current_step = "pillar_user"
    db.session.commit()

    login(client, "testcustomer", "custpass")
    resp = client.get(
        f"/assessments/{sample_assessment.id}?overview=1",
        follow_redirects=False,
    )
    assert resp.status_code == 200


def test_workspace_no_redirect_without_step(client, db, sample_assessment):
    assert sample_assessment.current_step is None

    login(client, "testcustomer", "custpass")
    resp = client.get(
        f"/assessments/{sample_assessment.id}",
        follow_redirects=False,
    )
    assert resp.status_code == 200


# ---------------------------------------------------------------------------
# Customer sensitive terms (#3)
# ---------------------------------------------------------------------------

def test_add_terms_creates_sensitive_term_rows(client, db, sample_assessment):
    login(client, "testcustomer", "custpass")
    resp = client.post(
        f"/assessments/{sample_assessment.id}/terms",
        data={"terms": "Acme Corp\ndb01.internal"},
        follow_redirects=True,
    )
    assert resp.status_code == 200

    terms = SensitiveTerm.query.filter_by(
        assessment_id=sample_assessment.id, source="user_added", is_active=True
    ).all()
    term_values = {t.term for t in terms}
    assert "Acme Corp" in term_values
    assert "db01.internal" in term_values


def test_add_terms_assigns_custom_tokens(client, db, sample_assessment):
    login(client, "testcustomer", "custpass")
    client.post(
        f"/assessments/{sample_assessment.id}/terms",
        data={"terms": "MyCorp"},
        follow_redirects=True,
    )
    term = SensitiveTerm.query.filter_by(
        assessment_id=sample_assessment.id, term="MyCorp"
    ).first()
    assert term is not None
    assert term.replacement_token.startswith("[CUSTOM_")


def test_add_terms_skips_duplicates(client, db, sample_assessment):
    login(client, "testcustomer", "custpass")
    # Add once
    client.post(
        f"/assessments/{sample_assessment.id}/terms",
        data={"terms": "UniqueOrg"},
        follow_redirects=True,
    )
    # Add again
    client.post(
        f"/assessments/{sample_assessment.id}/terms",
        data={"terms": "UniqueOrg"},
        follow_redirects=True,
    )
    count = SensitiveTerm.query.filter_by(
        assessment_id=sample_assessment.id, term="UniqueOrg", is_active=True
    ).count()
    assert count == 1


def test_add_terms_requires_editable_assessment(client, db, sample_assessment):
    sample_assessment.status = "awaiting_review"
    db.session.commit()

    login(client, "testcustomer", "custpass")
    resp = client.post(
        f"/assessments/{sample_assessment.id}/terms",
        data={"terms": "Should Not Add"},
        follow_redirects=True,
    )
    assert resp.status_code == 200
    count = SensitiveTerm.query.filter_by(
        assessment_id=sample_assessment.id, term="Should Not Add"
    ).count()
    assert count == 0


def test_workspace_shows_user_terms(client, db, sample_assessment):
    # Pre-seed a user term
    term = SensitiveTerm(
        assessment_id=sample_assessment.id,
        term="Secret Corp",
        replacement_token="[CUSTOM_1]",
        source="user_added",
        is_active=True,
    )
    db.session.add(term)
    db.session.commit()

    login(client, "testcustomer", "custpass")
    resp = client.get(
        f"/assessments/{sample_assessment.id}?overview=1",
    )
    assert b"Secret Corp" in resp.data


# ---------------------------------------------------------------------------
# Tool mapping — GET page (#6)
# ---------------------------------------------------------------------------

def test_tool_mapping_page_requires_admin(client, db, sample_assessment):
    login(client, "testcustomer", "custpass")
    tool = _make_tool(db, sample_assessment.id)
    resp = client.get(
        f"/admin/assessments/{sample_assessment.id}/inventory/{tool.id}/mapping",
        follow_redirects=False,
    )
    # Redirect to admin unlock
    assert resp.status_code == 302


def test_tool_mapping_page_loads(client, db, sample_assessment):
    login(client, "testcustomer", "custpass")
    _unlock_admin(client)
    tool = _make_tool(db, sample_assessment.id)
    resp = client.get(
        f"/admin/assessments/{sample_assessment.id}/inventory/{tool.id}/mapping",
    )
    assert resp.status_code == 200
    assert b"Activity Mapping" in resp.data
    assert tool.name.encode() in resp.data


# ---------------------------------------------------------------------------
# Tool mapping — suggest (#6)
# ---------------------------------------------------------------------------

def test_suggest_creates_ai_suggested_rows(client, db, sample_assessment):
    login(client, "testcustomer", "custpass")
    _unlock_admin(client)
    tool = _make_tool(db, sample_assessment.id)

    mock_suggestions = [
        {"activity_id": "1.1.1", "confidence": "high", "rationale": "EDR covers endpoint."},
    ]

    with patch(
        "app.routes.admin.suggest_mappings",
        return_value=(mock_suggestions, None, "prompt", "raw_response", "claude-sonnet-4-6"),
    ):
        resp = client.post(
            f"/admin/assessments/{sample_assessment.id}/inventory/{tool.id}/mapping/suggest",
            follow_redirects=True,
        )

    assert resp.status_code == 200
    mappings = ToolActivityMapping.query.filter_by(tool_id=tool.id, source="ai_suggested").all()
    assert len(mappings) == 1
    assert mappings[0].activity_id == "1.1.1"
    assert mappings[0].ai_confidence == "high"


def test_suggest_logs_to_suggestions_log(client, db, sample_assessment):
    login(client, "testcustomer", "custpass")
    _unlock_admin(client)
    tool = _make_tool(db, sample_assessment.id)

    with patch(
        "app.routes.admin.suggest_mappings",
        return_value=([], None, "my prompt", "my response", "claude-sonnet-4-6"),
    ):
        client.post(
            f"/admin/assessments/{sample_assessment.id}/inventory/{tool.id}/mapping/suggest",
            follow_redirects=True,
        )

    log = MappingSuggestionsLog.query.filter_by(tool_id=tool.id).first()
    assert log is not None
    assert log.request_payload == "my prompt"
    assert log.model_used == "claude-sonnet-4-6"


def test_suggest_replaces_old_ai_suggestions(client, db, sample_assessment):
    login(client, "testcustomer", "custpass")
    _unlock_admin(client)
    tool = _make_tool(db, sample_assessment.id)

    # Pre-seed old AI suggestion
    old = ToolActivityMapping(
        tool_id=tool.id, activity_id="old.1.1", source="ai_suggested"
    )
    db.session.add(old)
    db.session.commit()

    new_suggestions = [
        {"activity_id": "1.1.1", "confidence": "medium", "rationale": "New suggestion."}
    ]
    with patch(
        "app.routes.admin.suggest_mappings",
        return_value=(new_suggestions, None, "p", "r", "claude-sonnet-4-6"),
    ):
        client.post(
            f"/admin/assessments/{sample_assessment.id}/inventory/{tool.id}/mapping/suggest",
            follow_redirects=True,
        )

    all_ai = ToolActivityMapping.query.filter_by(tool_id=tool.id, source="ai_suggested").all()
    ids = {m.activity_id for m in all_ai}
    assert "old.1.1" not in ids
    assert "1.1.1" in ids


def test_suggest_error_shows_warning(client, db, sample_assessment):
    login(client, "testcustomer", "custpass")
    _unlock_admin(client)
    tool = _make_tool(db, sample_assessment.id)

    with patch(
        "app.routes.admin.suggest_mappings",
        return_value=([], "API call failed: timeout"),
    ):
        resp = client.post(
            f"/admin/assessments/{sample_assessment.id}/inventory/{tool.id}/mapping/suggest",
            follow_redirects=True,
        )

    assert resp.status_code == 200
    assert b"map manually" in resp.data or b"unavailable" in resp.data


# ---------------------------------------------------------------------------
# Tool mapping — finalize (#6)
# ---------------------------------------------------------------------------

def test_finalize_sets_tool_status_active(client, db, sample_assessment):
    login(client, "testcustomer", "custpass")
    _unlock_admin(client)
    tool = _make_tool(db, sample_assessment.id)

    # Pre-seed an AI suggestion
    ai_m = ToolActivityMapping(
        tool_id=tool.id, activity_id="1.1.1", source="ai_suggested", ai_confidence="high"
    )
    db.session.add(ai_m)
    db.session.commit()

    resp = client.post(
        f"/admin/assessments/{sample_assessment.id}/inventory/{tool.id}/mapping/finalize",
        data={"activity_ids": ["1.1.1"]},
        follow_redirects=True,
    )
    assert resp.status_code == 200

    db.session.refresh(tool)
    assert tool.mapping_status == "active"
    assert tool.mappings_finalized_at is not None


def test_finalize_creates_admin_confirmed_row_for_ai_suggested(client, db, sample_assessment):
    login(client, "testcustomer", "custpass")
    _unlock_admin(client)
    tool = _make_tool(db, sample_assessment.id)

    ai_m = ToolActivityMapping(
        tool_id=tool.id, activity_id="1.1.1", source="ai_suggested", ai_confidence="high"
    )
    db.session.add(ai_m)
    db.session.commit()

    client.post(
        f"/admin/assessments/{sample_assessment.id}/inventory/{tool.id}/mapping/finalize",
        data={"activity_ids": ["1.1.1"]},
        follow_redirects=True,
    )

    confirmed = ToolActivityMapping.query.filter_by(
        tool_id=tool.id, activity_id="1.1.1"
    ).first()
    assert confirmed is not None
    assert confirmed.source == "admin_confirmed"


def test_finalize_creates_admin_added_row_for_manual_selection(client, db, sample_assessment):
    login(client, "testcustomer", "custpass")
    _unlock_admin(client)
    tool = _make_tool(db, sample_assessment.id)
    db.session.commit()

    # No AI suggestions for this activity — admin adding manually
    resp = client.post(
        f"/admin/assessments/{sample_assessment.id}/inventory/{tool.id}/mapping/finalize",
        data={"activity_ids": ["1.1.1"]},
        follow_redirects=True,
    )
    assert resp.status_code == 200

    added = ToolActivityMapping.query.filter_by(
        tool_id=tool.id, activity_id="1.1.1"
    ).first()
    assert added is not None
    assert added.source == "admin_added"


def test_finalize_rejects_empty_selection(client, db, sample_assessment):
    login(client, "testcustomer", "custpass")
    _unlock_admin(client)
    tool = _make_tool(db, sample_assessment.id)

    resp = client.post(
        f"/admin/assessments/{sample_assessment.id}/inventory/{tool.id}/mapping/finalize",
        data={},
        follow_redirects=True,
    )
    assert resp.status_code == 200
    assert b"Select at least one" in resp.data

    db.session.refresh(tool)
    assert tool.mapping_status == "pending_review"


def test_finalize_logs_mapping_change_on_re_finalize(client, db, sample_assessment):
    login(client, "testcustomer", "custpass")
    _unlock_admin(client)
    tool = _make_tool(db, sample_assessment.id)

    # First finalization
    ai_m = ToolActivityMapping(
        tool_id=tool.id, activity_id="1.1.1", source="ai_suggested"
    )
    db.session.add(ai_m)
    db.session.commit()
    client.post(
        f"/admin/assessments/{sample_assessment.id}/inventory/{tool.id}/mapping/finalize",
        data={"activity_ids": ["1.1.1"]},
        follow_redirects=True,
    )

    # Second finalization (different selection) — should log a MappingChange
    ai_m2 = ToolActivityMapping(
        tool_id=tool.id, activity_id="1.1.2", source="ai_suggested"
    )
    db.session.add(ai_m2)
    db.session.commit()
    client.post(
        f"/admin/assessments/{sample_assessment.id}/inventory/{tool.id}/mapping/finalize",
        data={"activity_ids": ["1.1.2"]},
        follow_redirects=True,
    )

    changes = MappingChange.query.filter_by(tool_id=tool.id).all()
    assert len(changes) == 1
    after = json.loads(changes[0].after_state)
    assert "1.1.2" in after


# ---------------------------------------------------------------------------
# Excel — Gap Register Priority / Related Tools (#1)
# ---------------------------------------------------------------------------

def test_excel_gap_register_has_priority_column(db, sample_assessment):
    from app.services.excel_service import build_customer_excel
    from openpyxl import load_workbook
    import io

    xlsx = build_customer_excel(sample_assessment)
    wb = load_workbook(io.BytesIO(xlsx))

    # Find Gap Register sheet
    sheet_names = [s.lower() for s in wb.sheetnames]
    gap_sheet = next(
        (wb[name] for name in wb.sheetnames if "gap" in name.lower()), None
    )
    assert gap_sheet is not None, f"No gap sheet found in {wb.sheetnames}"

    headers = [cell.value for cell in gap_sheet[1] if cell.value]
    assert any("priority" in str(h).lower() for h in headers), f"Headers: {headers}"


def test_excel_gap_register_has_related_tools_column(db, sample_assessment):
    from app.services.excel_service import build_customer_excel
    from openpyxl import load_workbook
    import io

    xlsx = build_customer_excel(sample_assessment)
    wb = load_workbook(io.BytesIO(xlsx))

    gap_sheet = next(
        (wb[name] for name in wb.sheetnames if "gap" in name.lower()), None
    )
    assert gap_sheet is not None

    headers = [cell.value for cell in gap_sheet[1] if cell.value]
    assert any("tool" in str(h).lower() for h in headers), f"Headers: {headers}"


# ---------------------------------------------------------------------------
# Excel — Top 5 callout in Executive Summary (#2)
# ---------------------------------------------------------------------------

def test_excel_executive_summary_has_top5_callout(db, sample_assessment):
    from app.services.excel_service import build_customer_excel
    from openpyxl import load_workbook
    import io

    xlsx = build_customer_excel(sample_assessment)
    wb = load_workbook(io.BytesIO(xlsx))

    exec_sheet = next(
        (wb[name] for name in wb.sheetnames if "exec" in name.lower() or "summary" in name.lower()),
        None,
    )
    assert exec_sheet is not None, f"No exec summary sheet in {wb.sheetnames}"

    all_text = " ".join(
        str(cell.value) for row in exec_sheet.iter_rows() for cell in row if cell.value
    )
    assert "top 5" in all_text.lower() or "priority gap" in all_text.lower(), \
        f"Top 5 callout not found. Partial text: {all_text[:500]}"


# ---------------------------------------------------------------------------
# SharePoint README.txt (#7)
# ---------------------------------------------------------------------------

def test_upload_assessment_outputs_includes_readme():
    from app.services.sharepoint_service import upload_assessment_outputs
    from unittest.mock import MagicMock, patch

    client = MagicMock()
    client.ensure_folder = MagicMock()
    client.upload_file = MagicMock(return_value="https://sp/file")

    upload_assessment_outputs(
        client=client,
        assessment_id="a1",
        org_name="Acme",
        finalized_at=datetime.now(timezone.utc),
        customer_xlsx=b"cust",
        consultant_xlsx=b"cons",
        responses_json='{"data": 1}',
        ai_call_log_rows=[{"col": "val"}],
        audit_log_rows=[{"col": "val"}],
    )

    # Should be 6 uploads now (README + 5 original files)
    assert client.upload_file.call_count == 6

    # Find the README call
    readme_calls = [
        call for call in client.upload_file.call_args_list
        if "README.txt" in str(call)
    ]
    assert len(readme_calls) == 1


def test_readme_content_includes_org_name():
    from app.services.sharepoint_service import _build_readme

    content = _build_readme("Acme Corp", "assess-123", datetime.now(timezone.utc))
    text = content.decode("utf-8")
    assert "Acme Corp" in text
    assert "assess-123" in text
    assert "customer_report.xlsx" in text
