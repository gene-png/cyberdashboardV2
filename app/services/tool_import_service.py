"""
AI-powered tool import: parse any uploaded file to extract tool inventory entries.
"""
import io
import csv
import json
import logging
import os

logger = logging.getLogger(__name__)

TOOL_CATEGORIES = [
    "EDR", "XDR", "SIEM", "SOAR", "PAM", "IAM", "MFA", "SSO",
    "Firewall", "NDR", "DLP", "CASB", "WAF", "Vulnerability Management",
    "Patch Management", "Email Security", "Web Proxy", "DNS Security",
    "Endpoint Protection", "Cloud Security", "ZTNA", "Microsegmentation",
    "Certificate Management", "Secrets Management", "MDM", "Other",
]


def extract_file_text(file_storage) -> str:
    """Extract text from a werkzeug FileStorage object."""
    filename = file_storage.filename or ""
    ext = os.path.splitext(filename)[1].lower()
    data = file_storage.read()
    file_storage.seek(0)

    try:
        if ext == ".pdf":
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(data))
            return "\n".join(p.extract_text() or "" for p in reader.pages)[:15000]
        if ext in (".docx",):
            from docx import Document
            doc = Document(io.BytesIO(data))
            return "\n".join(p.text for p in doc.paragraphs)[:15000]
        if ext in (".xlsx",):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            lines = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    line = "\t".join(str(c) if c is not None else "" for c in row)
                    if line.strip():
                        lines.append(line)
            return "\n".join(lines)[:15000]
        # Try text decode
        import chardet
        enc = chardet.detect(data[:10000]).get("encoding") or "utf-8"
        return data.decode(enc, errors="replace")[:15000]
    except Exception as exc:
        logger.warning("Tool import file extraction failed for %s: %s", filename, exc)
        return data.decode("utf-8", errors="replace")[:15000]


def parse_tools_with_ai(file_text: str, api_key: str, model: str) -> list[dict]:
    """
    Send file content to Claude and extract a list of security tools.
    Returns list of {name, vendor, category, notes} dicts (max 50).
    """
    if not api_key:
        return _try_csv_parse(file_text)

    categories_str = ", ".join(TOOL_CATEGORIES)
    prompt = (
        f"The following text contains information about security tools an organization uses.\n\n"
        f"Extract every distinct security tool you can identify and return ONLY a JSON array.\n"
        f"Each element must have exactly these keys: name, vendor, category, notes.\n"
        f"- name: the tool/product name (required, non-empty)\n"
        f"- vendor: the company that makes it (empty string if unknown)\n"
        f"- category: best-fit from this list: {categories_str} (use 'Other' if unsure)\n"
        f"- notes: any useful detail from the text (version, deployment scope) — keep under 100 chars\n"
        f"Return at most 50 tools. Respond with ONLY the JSON array, no markdown fences.\n\n"
        f"Text:\n{file_text[:10000]}"
    )

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        message = client.messages.create(
            model=model,
            max_tokens=2048,
            system="You are a data extraction assistant. Respond only with valid JSON.",
            messages=[{"role": "user", "content": prompt}],
        )
        raw = message.content[0].text if message.content else "[]"
        text = raw.strip()
        import re
        if text.startswith("```"):
            text = re.sub(r"```[a-z]*\n?", "", text).replace("```", "").strip()
        tools = json.loads(text)
        if not isinstance(tools, list):
            return []
        result = []
        for t in tools[:50]:
            if not isinstance(t, dict) or not t.get("name"):
                continue
            result.append({
                "name": str(t.get("name", ""))[:200],
                "vendor": str(t.get("vendor", ""))[:200],
                "category": str(t.get("category", "Other"))[:100],
                "notes": str(t.get("notes", ""))[:500],
            })
        return result
    except Exception as exc:
        logger.warning("AI tool parse failed: %s", exc)
        return _try_csv_parse(file_text)


def _try_csv_parse(text: str) -> list[dict]:
    """Fallback: try to parse as CSV with headers name/vendor/category/notes."""
    try:
        reader = csv.DictReader(io.StringIO(text))
        results = []
        for row in reader:
            name = (row.get("name") or row.get("Name") or row.get("Tool Name") or "").strip()
            if not name:
                continue
            results.append({
                "name": name[:200],
                "vendor": (row.get("vendor") or row.get("Vendor") or "").strip()[:200],
                "category": (row.get("category") or row.get("Category") or "Other").strip()[:100],
                "notes": (row.get("notes") or row.get("Notes") or "").strip()[:500],
            })
        return results[:50]
    except Exception:
        return []


def build_csv_template() -> str:
    """Return a sample CSV template string."""
    lines = [
        "name,vendor,category,notes",
        "Microsoft Defender for Endpoint,Microsoft,EDR,Deployed on all Windows endpoints",
        "CrowdStrike Falcon,CrowdStrike,EDR,",
        "Splunk Enterprise SIEM,Splunk,SIEM,On-prem deployment",
        "Okta,Okta,IAM,SSO and MFA for cloud apps",
        "Palo Alto NGFW,Palo Alto Networks,Firewall,Perimeter and east-west",
    ]
    return "\n".join(lines)
