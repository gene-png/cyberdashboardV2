"""
Evidence upload processing and AI-assisted current state suggestions.
"""
import os
import logging
from flask import current_app
from ..extensions import db
from ..models import Response, PillarEvidence
from .scrub_service import scrub, rehydrate

logger = logging.getLogger(__name__)

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}
TEXT_EXTENSIONS = {".txt", ".csv", ".json", ".md", ".log", ".xml", ".yaml", ".yml"}


def extract_text(file_path: str, original_filename: str) -> str:
    """Extract readable text from an uploaded file."""
    ext = os.path.splitext(original_filename)[1].lower()
    try:
        if ext == ".pdf":
            from pypdf import PdfReader
            reader = PdfReader(file_path)
            return "\n".join(page.extract_text() or "" for page in reader.pages)[:20000]
        if ext in (".docx", ".doc"):
            from docx import Document
            doc = Document(file_path)
            return "\n".join(p.text for p in doc.paragraphs)[:20000]
        if ext in (".xlsx", ".xls"):
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            lines = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    line = "\t".join(str(c) if c is not None else "" for c in row)
                    if line.strip():
                        lines.append(line)
            return "\n".join(lines)[:20000]
        if ext in TEXT_EXTENSIONS:
            import chardet
            with open(file_path, "rb") as f:
                raw = f.read(50000)
            enc = chardet.detect(raw).get("encoding") or "utf-8"
            return raw.decode(enc, errors="replace")[:20000]
        if ext in IMAGE_EXTENSIONS:
            return f"[IMAGE: {original_filename}]"  # Claude vision handled separately
    except Exception as exc:
        logger.warning("Text extraction failed for %s: %s", original_filename, exc)
    return ""


def suggest_states_from_evidence(
    assessment_id: str,
    pillar_id: str,
    pillar_name: str,
    activities: list[dict],
    framework_name: str,
    maturity_states: list[str],
    maturity_labels: dict,
) -> dict:
    """
    Call Claude to suggest current_state for each activity based on uploaded evidence.
    Returns dict: activity_id -> suggested_state (only fills where no existing state).
    """
    api_key = current_app.config.get("ANTHROPIC_API_KEY", "")
    model = current_app.config.get("ANTHROPIC_MODEL", "claude-sonnet-4-6")
    if not api_key:
        return {}

    evidences = PillarEvidence.query.filter_by(
        assessment_id=assessment_id, pillar_name=pillar_id
    ).all()
    if not evidences:
        return {}

    evidence_texts = []
    for ev in evidences:
        if ev.extracted_text and ev.extracted_text.startswith("[IMAGE:"):
            evidence_texts.append(f"[Uploaded image: {ev.original_filename}]")
        elif ev.extracted_text:
            evidence_texts.append(f"=== {ev.original_filename} ===\n{ev.extracted_text[:3000]}")

    combined = "\n\n".join(evidence_texts) if evidence_texts else "(no text extractable)"
    states_list = ", ".join(f"{s} ({maturity_labels.get(s, s)})" for s in maturity_states)

    activities_block = "\n".join(
        f"- {a['id']}: {a['name']}" for a in activities
    )

    prompt = (
        f"Framework: {framework_name}\n"
        f"Pillar: {pillar_name}\n\n"
        f"Activities to assess:\n{activities_block}\n\n"
        f"Evidence provided:\n{combined[:8000]}\n\n"
        f"Maturity states (ordered lowest to highest): {states_list}\n\n"
        f"Task: Based ONLY on the evidence above, suggest a current maturity state for each activity. "
        f"If the evidence does not support a higher state, use the lowest ({maturity_states[0]}). "
        f"Respond with ONLY a JSON object mapping activity_id to state value, e.g. "
        f'{{"{activities[0]["id"]}": "{maturity_states[0]}"}}. '
        f"Use exact state values from the list provided."
    )

    scrubbed = scrub(assessment_id, prompt)

    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        message = client.messages.create(
            model=model,
            max_tokens=512,
            system="You are a cybersecurity maturity assessor. Respond with only valid JSON.",
            messages=[{"role": "user", "content": scrubbed}],
        )
        raw = message.content[0].text if message.content else ""
        import json, re
        text = raw.strip()
        if text.startswith("```"):
            text = re.sub(r"```[a-z]*\n?", "", text).replace("```", "").strip()
        suggestions = json.loads(text)
        if isinstance(suggestions, dict):
            valid = {aid: sv for aid, sv in suggestions.items() if sv in maturity_states}
            return {aid: rehydrate(assessment_id, sv) if sv not in maturity_states else sv
                    for aid, sv in valid.items()}
    except Exception as exc:
        logger.warning("Evidence AI suggestion failed: %s", exc)
    return {}


def apply_initial_defaults(assessment_id: str, pillar_id: str, activities: list[dict]) -> int:
    """
    For activities with no evidence and no notes and no current_state, set to first state ("initial").
    Returns count of activities defaulted.
    """
    has_evidence = PillarEvidence.query.filter_by(
        assessment_id=assessment_id, pillar_name=pillar_id
    ).count() > 0
    if has_evidence:
        return 0
    count = 0
    for activity in activities:
        resp = Response.query.filter_by(
            assessment_id=assessment_id, activity_id=activity["id"]
        ).first()
        if resp and not resp.current_state_value and not resp.evidence_notes:
            resp.current_state_value = "initial"
            count += 1
        elif resp is None:
            resp = Response(
                assessment_id=assessment_id,
                pillar=pillar_id,
                activity_id=activity["id"],
                current_state_value="initial",
            )
            db.session.add(resp)
            count += 1
    if count:
        db.session.commit()
    return count
