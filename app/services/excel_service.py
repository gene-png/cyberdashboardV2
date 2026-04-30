"""Generates customer and consultant Excel reports using openpyxl."""
import io
import re
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .framework_loader import load_framework
from ..models import Assessment, Response, AdminScore, GapFinding, ToolInventory, AICallLog, AuditLog, ToolActivityMapping


# Color palette
RED = "00FF0000"
AMBER = "00FFA500"
GREEN = "0000AA00"
LIGHT_BLUE = "00DDEEFF"
DARK_BLUE = "00003366"
WHITE = "00FFFFFF"
LIGHT_GREY = "00F5F5F5"


def _header_style(ws, row, cols, text, bg=DARK_BLUE, fg=WHITE, bold=True):
    fill = PatternFill("solid", fgColor=bg)
    font = Font(color=fg, bold=bold)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
    ws.cell(row=row, column=1).value = text


def _gap_fill(gap_int: int) -> PatternFill | None:
    if gap_int <= 0:
        return PatternFill("solid", fgColor=GREEN)
    if gap_int == 1:
        return PatternFill("solid", fgColor=AMBER)
    return PatternFill("solid", fgColor=RED)


def _compute_pillar_stats(assessment: Assessment, framework: dict) -> list[dict]:
    maturity_order = framework["maturity_order"]
    maturity_states = framework["maturity_states"]
    responses_by_activity = {r.activity_id: r for r in assessment.responses}

    stats = []
    for pillar in framework["pillars"]:
        pillar_responses = [
            responses_by_activity.get(a["id"])
            for a in pillar["activities"]
            if responses_by_activity.get(a["id"])
        ]
        total = len(pillar["activities"])
        met = sum(
            1 for r in pillar_responses
            if r.current_state_value and r.target_state_value
            and maturity_order.get(r.current_state_value, 0) >= maturity_order.get(r.target_state_value, 0)
        )
        partial = sum(
            1 for r in pillar_responses
            if r.current_state_value and r.target_state_value
            and maturity_order.get(r.current_state_value, 0) == maturity_order.get(r.target_state_value, 0) - 1
        )
        not_met = sum(
            1 for r in pillar_responses
            if r.current_state_value and r.target_state_value
            and maturity_order.get(r.current_state_value, 0) < maturity_order.get(r.target_state_value, 0) - 1
        )
        # Score = average current / max possible (0-100)
        if pillar_responses:
            max_val = len(maturity_states) - 1
            cur_avg = sum(maturity_order.get(r.current_state_value, 0) for r in pillar_responses) / len(pillar_responses)
            tgt_avg = sum(maturity_order.get(r.target_state_value, 0) for r in pillar_responses if r.target_state_value) / max(1, len([r for r in pillar_responses if r.target_state_value]))
            current_score = round((cur_avg / max_val) * 100, 1)
            target_score = round((tgt_avg / max_val) * 100, 1)
        else:
            current_score = target_score = 0.0

        stats.append({
            "pillar_id": pillar["id"],
            "pillar_name": pillar["name"],
            "total": total,
            "met": met,
            "gap_small": partial,   # gap of exactly 1 maturity step
            "gap_large": not_met,   # gap of 2+ maturity steps
            "current_score": current_score,
            "target_score": target_score,
            "gap": round(target_score - current_score, 1),
        })
    return stats


_SEVERITY_PRIORITY = {"critical": 1, "high": 2, "medium": 3, "low": 4}


def _build_activity_to_tools_map(assessment: Assessment) -> dict[str, list[str]]:
    """Return {activity_id: [tool_name, ...]} for all active tool mappings."""
    result: dict[str, list[str]] = {}
    for tool in assessment.tool_inventory:
        if tool.mapping_status != "active":
            continue
        for m in tool.activity_mappings:
            if m.source in ("admin_confirmed", "admin_added"):
                result.setdefault(m.activity_id, []).append(tool.name)
    return result


def build_customer_excel(assessment: Assessment) -> bytes:
    framework = load_framework(assessment.framework)
    maturity_order = framework["maturity_order"]
    wb = Workbook()

    responses_by_activity = {r.activity_id: r for r in assessment.responses}
    findings_by_activity = {f.activity_id: f for f in assessment.gap_findings}
    activity_to_tools = _build_activity_to_tools_map(assessment)

    # ---- Sheet 1: Executive Summary ----
    ws = wb.active
    ws.title = "Executive Summary"

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10

    ws["A1"] = "Zero Trust Maturity Assessment"
    ws["A1"].font = Font(bold=True, size=16, color=DARK_BLUE)
    ws["A2"] = f"Organization: {assessment.customer_org}"
    ws["A3"] = f"Framework: {framework['name']}"
    ws["A4"] = f"Assessment Date: {assessment.finalized_at.strftime('%Y-%m-%d') if assessment.finalized_at else datetime.now(timezone.utc).strftime('%Y-%m-%d')}"
    ws["A5"] = f"Status: {assessment.status.replace('_', ' ').title()}"

    ws.append([])
    header_row = ws.max_row + 1
    headers = ["Pillar", "Current Score", "Target Score", "Gap %", "Large Gap", "Small Gap", "Met"]
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = Alignment(horizontal="center")

    pillar_stats = _compute_pillar_stats(assessment, framework)
    for stat in pillar_stats:
        row = [
            stat["pillar_name"],
            stat["current_score"],
            stat["target_score"],
            stat["gap"],
            stat["gap_large"],
            stat["gap_small"],
            stat["met"],
        ]
        ws.append(row)
        data_row = ws.max_row
        gap_fill = _gap_fill(int(stat["gap"]))
        if gap_fill:
            ws.cell(row=data_row, column=4).fill = gap_fill
        for col in range(2, 8):
            ws.cell(row=data_row, column=col).alignment = Alignment(horizontal="center")

    # ---- Top 5 Priority Gaps callout ----
    all_gaps = sorted(
        [f for f in assessment.gap_findings if f.severity],
        key=lambda f: _SEVERITY_PRIORITY.get(f.severity, 99),
    )[:5]
    ws.append([])
    callout_row = ws.max_row + 1
    ws.append(["Top 5 Priority Gaps"])
    ws.cell(row=callout_row, column=1).font = Font(bold=True, size=12, color=DARK_BLUE)
    if all_gaps:
        ws.append(["Pillar", "Activity", "Gap Status"])
        for col in range(1, 4):
            cell = ws.cell(row=callout_row + 1, column=col)
            cell.font = Font(bold=True, color=WHITE)
            cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        for f in all_gaps:
            pillar_name = next(
                (p["name"] for p in framework["pillars"]
                 for a in p["activities"] if a["id"] == f.activity_id),
                f.pillar,
            )
            activity_name = next(
                (a["name"] for p in framework["pillars"]
                 for a in p["activities"] if a["id"] == f.activity_id),
                f.activity_id,
            )
            ws.append([pillar_name, activity_name, f.severity.title()])
    else:
        ws.append(["No priority gaps identified yet."])

    # ---- Sheet 2: Gap Register ----
    ws2 = wb.create_sheet("Gap Register")
    gap_headers = [
        "Pillar", "Activity ID", "Activity", "Current State", "Target State",
        "Gap Status", "Priority", "AI Remediation Guidance", "Related Tools",
    ]
    ws2.append(gap_headers)
    for col, h in enumerate(gap_headers, 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 38
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 14
    ws2.column_dimensions["F"].width = 12
    ws2.column_dimensions["G"].width = 10
    ws2.column_dimensions["H"].width = 55
    ws2.column_dimensions["I"].width = 30

    for pillar in framework["pillars"]:
        for activity in pillar["activities"]:
            resp = responses_by_activity.get(activity["id"])
            if not resp:
                continue
            if not resp.current_state_value or not resp.target_state_value:
                continue
            if maturity_order.get(resp.current_state_value, 0) >= maturity_order.get(resp.target_state_value, 0):
                continue
            finding = findings_by_activity.get(activity["id"])
            severity = finding.severity if finding and finding.severity else None
            priority = _SEVERITY_PRIORITY.get(severity, "") if severity else ""
            guidance = finding.rehydrated_response if finding and finding.rehydrated_response else "Pending AI analysis"
            tools_str = " | ".join(activity_to_tools.get(activity["id"], [])) or "—"
            row = [
                pillar["name"],
                activity["id"],
                activity["name"],
                framework["maturity_labels"].get(resp.current_state_value, resp.current_state_value),
                framework["maturity_labels"].get(resp.target_state_value, resp.target_state_value),
                severity.title() if severity else "—",
                priority,
                guidance,
                tools_str,
            ]
            ws2.append(row)
            last_row = ws2.max_row
            ws2.cell(row=last_row, column=8).alignment = Alignment(wrap_text=True)
            ws2.row_dimensions[last_row].height = 80

    # ---- Per-pillar sheets ----
    _invalid_chars = re.compile(r'[\\/*?\[\]:]')
    for pillar in framework["pillars"]:
        safe_title = _invalid_chars.sub("_", pillar["name"])[:31]
        wsp = wb.create_sheet(safe_title)
        pillar_headers = ["Activity ID", "Activity", "Current State", "Target State", "Evidence Notes", "AI Guidance"]
        wsp.append(pillar_headers)
        for col, h in enumerate(pillar_headers, 1):
            cell = wsp.cell(row=1, column=col)
            cell.font = Font(bold=True, color=WHITE)
            cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        wsp.column_dimensions["A"].width = 22
        wsp.column_dimensions["B"].width = 40
        wsp.column_dimensions["C"].width = 14
        wsp.column_dimensions["D"].width = 14
        wsp.column_dimensions["E"].width = 40
        wsp.column_dimensions["F"].width = 60

        for activity in pillar["activities"]:
            resp = responses_by_activity.get(activity["id"])
            finding = findings_by_activity.get(activity["id"])
            row = [
                activity["id"],
                activity["name"],
                framework["maturity_labels"].get(resp.current_state_value, "") if resp and resp.current_state_value else "—",
                framework["maturity_labels"].get(resp.target_state_value, "") if resp and resp.target_state_value else "—",
                resp.evidence_notes or "" if resp else "",
                finding.rehydrated_response or "" if finding else "",
            ]
            wsp.append(row)
            last_row = wsp.max_row
            for col in [5, 6]:
                wsp.cell(row=last_row, column=col).alignment = Alignment(wrap_text=True)
            wsp.row_dimensions[last_row].height = 60

    # ---- Tool Inventory Mapping sheet ----
    _build_tool_inventory_mapping_sheet(wb, assessment, framework, activity_to_tools)

    # ---- Methodology sheet ----
    wsm = wb.create_sheet("Methodology")
    wsm["A1"] = "Assessment Methodology"
    wsm["A1"].font = Font(bold=True, size=14)
    wsm["A3"] = f"Framework: {framework['name']} v{framework['version']}"
    wsm["A5"] = "Scoring Rubric"
    wsm["A5"].font = Font(bold=True)
    for i, (state, label) in enumerate(framework["maturity_labels"].items(), 6):
        wsm[f"A{i}"] = f"  {label}: {state.replace('_', ' ').title()}"
    wsm["A12"] = "AI Guidance: Where shown, remediation guidance was generated by an AI model using scrubbed assessment data. All sensitive identifiers were replaced with tokens before transmission."
    wsm["A12"].alignment = Alignment(wrap_text=True)
    wsm.column_dimensions["A"].width = 80
    wsm.row_dimensions[12].height = 60

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_tool_inventory_mapping_sheet(
    wb, assessment: Assessment, framework: dict, activity_to_tools: dict
) -> None:
    """Build the Tool Inventory Mapping sheet with activity linkage and analysis panels."""
    from flask import current_app
    try:
        redundancy_threshold = current_app.config.get("REDUNDANCY_THRESHOLD", 3)
        underutil_min = current_app.config.get("TOOL_MIN_ACTIVITIES", 2)
    except RuntimeError:
        redundancy_threshold = 3
        underutil_min = 2

    active_tools = [t for t in assessment.tool_inventory if t.mapping_status == "active"]
    all_activities = {
        a["id"]: a["name"]
        for p in framework["pillars"]
        for a in p["activities"]
    }

    wst = wb.create_sheet("Tool Inventory Mapping")

    # Tool list section
    _header_style(wst, wst.max_row + 1, 5, "Tool Inventory", bg=DARK_BLUE)
    wst.append(["Tool Name", "Vendor", "Category", "Mapping Status", "Mapped Activities"])
    hrow = wst.max_row
    for col in range(1, 6):
        cell = wst.cell(row=hrow, column=col)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)

    wst.column_dimensions["A"].width = 28
    wst.column_dimensions["B"].width = 18
    wst.column_dimensions["C"].width = 16
    wst.column_dimensions["D"].width = 16
    wst.column_dimensions["E"].width = 55

    for tool in assessment.tool_inventory:
        confirmed = [
            m.activity_id for m in tool.activity_mappings
            if m.source in ("admin_confirmed", "admin_added")
        ]
        activity_names = [all_activities.get(aid, aid) for aid in confirmed]
        wst.append([
            tool.name,
            tool.vendor or "",
            tool.category or "",
            tool.mapping_status.replace("_", " ").title(),
            " | ".join(activity_names) if activity_names else "—",
        ])
        last_row = wst.max_row
        wst.cell(row=last_row, column=5).alignment = Alignment(wrap_text=True)
        wst.row_dimensions[last_row].height = 40

    # Redundancy analysis
    wst.append([])
    _header_style(wst, wst.max_row + 1, 3, f"Redundancy — Activities with {redundancy_threshold}+ Tools", bg=AMBER, fg="00000000")
    wst.append(["Activity", "Tools Mapped", "Tool Names"])
    hrow2 = wst.max_row
    for col in range(1, 4):
        cell = wst.cell(row=hrow2, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)

    redundant = {
        aid: tools for aid, tools in activity_to_tools.items()
        if len(tools) >= redundancy_threshold
    }
    for aid, tools in sorted(redundant.items(), key=lambda x: -len(x[1])):
        wst.append([all_activities.get(aid, aid), len(tools), " | ".join(tools)])

    if not redundant:
        wst.append(["No redundant coverage detected.", "", ""])

    # Underutilization analysis
    wst.append([])
    _header_style(
        wst, wst.max_row + 1, 3,
        f"Underutilization — Tools with fewer than {underutil_min} mapped activities",
        bg=AMBER, fg="00000000",
    )
    wst.append(["Tool Name", "Mapped Activities", "Status"])
    hrow3 = wst.max_row
    for col in range(1, 4):
        cell = wst.cell(row=hrow3, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)

    underutil = [
        t for t in active_tools
        if len([m for m in t.activity_mappings if m.source in ("admin_confirmed", "admin_added")]) < underutil_min
    ]
    for tool in sorted(underutil, key=lambda t: len(t.activity_mappings)):
        mapped_count = len([m for m in tool.activity_mappings if m.source in ("admin_confirmed", "admin_added")])
        wst.append([tool.name, mapped_count, "0 mappings — review tool scope" if mapped_count == 0 else f"Only {mapped_count} mapping(s)"])

    if not underutil:
        wst.append(["All active tools have adequate activity coverage.", "", ""])


def build_consultant_excel(assessment: Assessment) -> bytes:
    """Builds consultant report: everything in customer report plus admin notes, AI call log, audit log."""
    # Start with customer report content
    framework = load_framework(assessment.framework)
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # Re-use customer report in a temp workbook and copy sheets
    customer_bytes = build_customer_excel(assessment)
    import io as _io
    from openpyxl import load_workbook as _lw
    customer_wb = _lw(filename=_io.BytesIO(customer_bytes))
    for title in customer_wb.sheetnames:
        ws_src = customer_wb[title]
        ws_dst = wb.create_sheet(title)
        for row in ws_src.iter_rows():
            for cell in row:
                new_cell = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    from copy import copy as _copy
                    new_cell.font = _copy(cell.font)
                    new_cell.fill = _copy(cell.fill)
                    new_cell.alignment = _copy(cell.alignment)
        for col_letter, dim in ws_src.column_dimensions.items():
            ws_dst.column_dimensions[col_letter].width = dim.width
        for row_num, dim in ws_src.row_dimensions.items():
            ws_dst.row_dimensions[row_num].height = dim.height

    # ---- Admin Notes sheet ----
    wsa = wb.create_sheet("Admin Notes")
    wsa.append(["Pillar", "Current Score", "Target Score", "Gap Summary", "Consultant Recommendation"])
    for col in range(1, 6):
        cell = wsa.cell(row=1, column=col)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    wsa.column_dimensions["A"].width = 20
    wsa.column_dimensions["B"].width = 14
    wsa.column_dimensions["C"].width = 14
    wsa.column_dimensions["D"].width = 50
    wsa.column_dimensions["E"].width = 50
    for score in assessment.admin_scores:
        wsa.append([
            score.pillar, score.current_score, score.target_score,
            score.gap_summary or "", score.consultant_recommendation or "",
        ])

    # ---- AI Call Log sheet ----
    wsc = wb.create_sheet("AI Call Log")
    wsc.append(["Timestamp", "Model", "Tokens In", "Tokens Out", "Duration ms", "Scrubbed Prompt", "Scrubbed Response"])
    for col in range(1, 8):
        cell = wsc.cell(row=1, column=col)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    wsc.column_dimensions["A"].width = 20
    wsc.column_dimensions["B"].width = 20
    wsc.column_dimensions["C"].width = 12
    wsc.column_dimensions["D"].width = 12
    wsc.column_dimensions["E"].width = 14
    wsc.column_dimensions["F"].width = 60
    wsc.column_dimensions["G"].width = 60
    for log in assessment.ai_call_logs:
        wsc.append([
            str(log.timestamp), log.model, log.tokens_in, log.tokens_out, log.duration_ms,
            log.request_body_scrubbed or "", log.response_body_scrubbed or "",
        ])

    # ---- Audit Log sheet ----
    wsl = wb.create_sheet("Audit Log")
    wsl.append(["Timestamp", "User ID", "Action", "Target Type", "Target ID", "Before", "After"])
    for col in range(1, 8):
        cell = wsl.cell(row=1, column=col)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    wsl.column_dimensions["A"].width = 20
    for log in assessment.audit_logs:
        wsl.append([
            str(log.timestamp), log.user_id, log.action,
            log.target_type, log.target_id, log.before_value or "", log.after_value or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
