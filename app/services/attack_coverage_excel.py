"""
MITRE ATT&CK Coverage Report — Excel generator.

Builds a 5-sheet workbook from pre-computed coverage data.
"""
import io
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .attack_mapper import classify_gap_status

# ---- Colours ----
DARK_BLUE = "003366"
WHITE = "FFFFFF"
LIGHT_GREY = "F5F5F5"
RED_FILL = "FFCCCC"
YELLOW_FILL = "FFF2CC"
GREEN_FILL = "CCFFCC"
ORANGE_FILL = "FFE0CC"

_GAP_FILL = {
    "None": RED_FILL,
    "Single Tool": YELLOW_FILL,
    "Detect Only": YELLOW_FILL,
    "Prevent Only": YELLOW_FILL,
    "Full": GREEN_FILL,
}

_GAP_SORT_ORDER = {"None": 0, "Single Tool": 1, "Detect Only": 2, "Prevent Only": 3, "Full": 4}


def _header_row(ws, headers: list[str], row: int = 1) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autofit(ws, min_width: int = 10, max_width: int = 60) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def _freeze(ws) -> None:
    ws.freeze_panes = ws.cell(row=2, column=1)


def compute_coverage_matrix(coverage_data: list[dict], techniques: list) -> dict:
    """
    Build a technique → coverage dict from coverage_data.

    coverage_data: list of {tool, activity_ids, results}
    techniques:    list of MitreTechnique

    Returns: dict[full_id → {technique, detect_tools, prevent_tools, respond_tools,
                              gap_status, rationale_summary}]
    """
    matrix = {}

    # Pre-index techniques by full_id
    for t in techniques:
        matrix[t.full_id] = {
            "technique": t,
            "detect_tools": [],
            "prevent_tools": [],
            "respond_tools": [],
            "gap_status": "None",
            "rationale_summary": "",
        }

    # Populate from coverage data
    for entry in coverage_data:
        tool = entry["tool"]
        for result in entry["results"]:
            tid = result["technique_id"]
            if tid not in matrix:
                continue
            ctype = result["coverage_type"]
            rationale = result.get("rationale", "")
            if ctype == "detect":
                matrix[tid]["detect_tools"].append(tool.name)
            elif ctype == "prevent":
                matrix[tid]["prevent_tools"].append(tool.name)
            elif ctype == "respond":
                matrix[tid]["respond_tools"].append(tool.name)
            if rationale:
                existing = matrix[tid]["rationale_summary"]
                matrix[tid]["rationale_summary"] = (
                    f"{existing}; {tool.name}: {rationale}" if existing
                    else f"{tool.name}: {rationale}"
                )

    # Compute gap status
    for tid, entry in matrix.items():
        entry["gap_status"] = classify_gap_status(
            entry["detect_tools"], entry["prevent_tools"], entry["respond_tools"]
        )

    return matrix


def _build_summary_sheet(wb: Workbook, matrix: dict, coverage_data: list[dict],
                          generated_at: datetime, model_used: str) -> None:
    ws = wb.create_sheet("Summary", 0)
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20

    title_font = Font(bold=True, size=14, color=DARK_BLUE)
    header_font = Font(bold=True, color=DARK_BLUE)

    # Title
    ws.cell(row=1, column=1, value="MITRE ATT&CK Coverage Report").font = title_font
    ws.cell(row=2, column=1, value=f"Generated: {generated_at.strftime('%Y-%m-%d %H:%M UTC')}")
    ws.cell(row=3, column=1, value=f"Model: {model_used}")
    ws.cell(row=4, column=1, value=f"Tools Analyzed: {len(coverage_data)}")

    # Stats
    total = len(matrix)
    covered = sum(1 for v in matrix.values() if v["gap_status"] != "None")
    detect_covered = sum(1 for v in matrix.values() if v["detect_tools"])
    prevent_covered = sum(1 for v in matrix.values() if v["prevent_tools"])
    full_coverage = sum(1 for v in matrix.values() if v["gap_status"] == "Full")
    gap_count = total - covered
    coverage_pct = (covered / total * 100) if total else 0

    rows = [
        ("", ""),
        ("Metric", "Value"),
        ("Total ATT&CK Techniques", total),
        ("Techniques Covered (any tool)", covered),
        ("Detect Coverage", detect_covered),
        ("Prevent Coverage", prevent_covered),
        ("Full Coverage (detect + prevent)", full_coverage),
        ("Uncovered Gaps", gap_count),
        ("Overall Coverage %", f"{coverage_pct:.1f}%"),
    ]

    for i, (label, value) in enumerate(rows, start=6):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)
        if label == "Metric":
            ws.cell(row=i, column=1).font = header_font
            ws.cell(row=i, column=2).font = header_font

    # Top 10 gap techniques (status = "None")
    none_gaps = sorted(
        [(tid, v) for tid, v in matrix.items() if v["gap_status"] == "None"],
        key=lambda x: (x[1]["technique"].tactic or "", x[0]),
    )[:10]

    row = len(rows) + 8
    ws.cell(row=row, column=1, value="Top 10 Uncovered Techniques").font = Font(bold=True, size=12, color=DARK_BLUE)
    row += 1
    ws.cell(row=row, column=1, value="Technique ID")
    ws.cell(row=row, column=2, value="Name")
    ws.cell(row=row, column=3, value="Tactic")
    for col in range(1, 4):
        c = ws.cell(row=row, column=col)
        c.font = Font(bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    row += 1
    for tid, entry in none_gaps:
        t = entry["technique"]
        ws.cell(row=row, column=1, value=tid)
        ws.cell(row=row, column=2, value=t.name)
        ws.cell(row=row, column=3, value=t.tactic or "")
        for col in range(1, 4):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=RED_FILL)
        row += 1

    ws.column_dimensions["C"].width = 30


def _build_coverage_matrix_sheet(wb: Workbook, matrix: dict) -> None:
    ws = wb.create_sheet("Coverage Matrix")
    headers = [
        "Tactic", "Technique ID", "Technique Name", "Sub-technique",
        "Detect Tools", "Prevent Tools", "Respond Tools",
        "Gap Status", "Rationale Summary", "MITRE URL",
    ]
    _header_row(ws, headers)
    _freeze(ws)

    # Sort by tactic then technique ID
    sorted_entries = sorted(
        matrix.items(),
        key=lambda x: (x[1]["technique"].tactic or "", x[0]),
    )

    for row_idx, (tid, entry) in enumerate(sorted_entries, start=2):
        t = entry["technique"]
        gap = entry["gap_status"]
        fill_color = _GAP_FILL.get(gap, LIGHT_GREY)
        fill = PatternFill("solid", fgColor=fill_color)

        values = [
            t.tactic or "",
            t.technique_id,
            t.name,
            t.sub_technique_id or "",
            ", ".join(entry["detect_tools"]),
            ", ".join(entry["prevent_tools"]),
            ", ".join(entry["respond_tools"]),
            gap,
            entry["rationale_summary"][:500] if entry["rationale_summary"] else "",
            t.url or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if gap in ("None", "Single Tool", "Detect Only", "Prevent Only"):
                cell.fill = fill
            elif gap == "Full":
                cell.fill = PatternFill("solid", fgColor=GREEN_FILL)

    _autofit(ws)


def _build_gaps_sheet(wb: Workbook, matrix: dict) -> None:
    ws = wb.create_sheet("Gaps")
    headers = [
        "Tactic", "Technique ID", "Technique Name", "Sub-technique",
        "Gap Status", "Detect Tools", "Prevent Tools", "Respond Tools", "MITRE URL",
    ]
    _header_row(ws, headers)
    _freeze(ws)

    # Only non-Full entries, sorted by gap severity then tactic
    gap_entries = [
        (tid, entry) for tid, entry in matrix.items()
        if entry["gap_status"] != "Full"
    ]
    gap_entries.sort(
        key=lambda x: (
            _GAP_SORT_ORDER.get(x[1]["gap_status"], 99),
            x[1]["technique"].tactic or "",
            x[0],
        )
    )

    for row_idx, (tid, entry) in enumerate(gap_entries, start=2):
        t = entry["technique"]
        gap = entry["gap_status"]
        fill = PatternFill("solid", fgColor=_GAP_FILL.get(gap, LIGHT_GREY))

        values = [
            t.tactic or "",
            t.technique_id,
            t.name,
            t.sub_technique_id or "",
            gap,
            ", ".join(entry["detect_tools"]),
            ", ".join(entry["prevent_tools"]),
            ", ".join(entry["respond_tools"]),
            t.url or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val).fill = fill

    _autofit(ws)


def _build_tool_coverage_sheet(wb: Workbook, coverage_data: list[dict], matrix: dict) -> None:
    ws = wb.create_sheet("Tool Coverage")
    headers = [
        "Tool", "Vendor", "Category",
        "# Detect", "# Prevent", "# Respond", "# Total Techniques",
        "Coverage % of ATT&CK Enterprise",
    ]
    _header_row(ws, headers)
    _freeze(ws)

    total_techniques = len(matrix)

    for row_idx, entry in enumerate(coverage_data, start=2):
        tool = entry["tool"]
        results = entry["results"]
        detect = sum(1 for r in results if r["coverage_type"] == "detect")
        prevent = sum(1 for r in results if r["coverage_type"] == "prevent")
        respond = sum(1 for r in results if r["coverage_type"] == "respond")
        total = len({r["technique_id"] for r in results})
        pct = f"{total / total_techniques * 100:.1f}%" if total_techniques else "0.0%"

        values = [
            tool.name, tool.vendor or "", tool.category or "",
            detect, prevent, respond, total, pct,
        ]
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    _autofit(ws)


def _build_methodology_sheet(
    wb: Workbook,
    generated_at: datetime,
    model_used: str,
    coverage_data: list[dict],
    excluded_tools: list[str],
) -> None:
    ws = wb.create_sheet("Methodology")
    ws.column_dimensions["A"].width = 100

    lines = [
        ("MITRE ATT&CK Coverage Report — Methodology", True, 14),
        ("", False, 11),
        ("How Mappings Were Generated", True, 12),
        ("Each tool with finalized activity mappings was analyzed by an AI model that reviewed the "
         "tool's name, vendor, category, and description, along with its confirmed Zero Trust "
         "framework activity assignments. The model was asked to identify which MITRE ATT&CK "
         "Enterprise techniques the tool covers and classify each as detect, prevent, or respond.", False, 11),
        ("", False, 11),
        ("Gap Definitions", True, 12),
        ("  Full:         Tool has both detect AND prevent coverage for the technique.", False, 11),
        ("  Detect Only:  Tool(s) can detect/alert on the technique but cannot prevent it.", False, 11),
        ("  Prevent Only: Tool(s) can prevent the technique but lack detection capability.", False, 11),
        ("  Single Tool:  Exactly one tool covers the technique (any coverage type) — single point of failure.", False, 11),
        ("  None:         No tool in the inventory covers this technique.", False, 11),
        ("", False, 11),
        ("Coverage Types", True, 12),
        ("  detect  — The tool detects, alerts, or logs activity related to the technique.", False, 11),
        ("  prevent — The tool actively blocks, prevents, or mitigates the technique.", False, 11),
        ("  respond — The tool supports investigation or remediation after the technique is executed.", False, 11),
        ("", False, 11),
        ("Generation Details", True, 12),
        (f"  Generated:    {generated_at.strftime('%Y-%m-%d %H:%M UTC')}", False, 11),
        (f"  Model used:   {model_used}", False, 11),
        (f"  Tools analyzed: {len(coverage_data)}", False, 11),
    ]

    if excluded_tools:
        lines.append(("", False, 11))
        lines.append(("Excluded Tools (mapping_status = pending_review):", True, 11))
        for name in excluded_tools:
            lines.append((f"  {name}", False, 11))

    lines += [
        ("", False, 11),
        ("Important Notes", True, 12),
        ("  - AI-generated mappings are based on tool metadata and may not reflect every deployment configuration.", False, 11),
        ("  - Results should be reviewed by a qualified security analyst.", False, 11),
        ("  - Coverage does not imply full protection — a 'prevent' mapping means the tool is capable of preventing", False, 11),
        ("    the technique under the right configuration, not that it is necessarily configured to do so.", False, 11),
        ("  - MITRE ATT&CK Enterprise dataset version corresponds to the seeded data in the assessment platform.", False, 11),
    ]

    for row_idx, (text, bold, size) in enumerate(lines, start=1):
        cell = ws.cell(row=row_idx, column=1, value=text)
        cell.font = Font(bold=bold, size=size, color=DARK_BLUE if bold else "000000")
        cell.alignment = Alignment(wrap_text=True)


def build_attack_coverage_excel(
    coverage_data: list[dict],
    techniques: list,
    generated_at: datetime,
    model_used: str,
    excluded_tool_names: list[str] | None = None,
) -> bytes:
    """
    Build the 5-sheet ATT&CK coverage workbook.

    coverage_data: list of {tool, activity_ids, results}
    techniques: list of MitreTechnique objects
    generated_at: report generation timestamp
    model_used: model identifier string
    excluded_tool_names: names of tools excluded from the report

    Returns: xlsx file bytes
    """
    wb = Workbook()
    # Remove default empty sheet
    if wb.active:
        wb.remove(wb.active)

    matrix = compute_coverage_matrix(coverage_data, techniques)

    _build_summary_sheet(wb, matrix, coverage_data, generated_at, model_used)
    _build_coverage_matrix_sheet(wb, matrix)
    _build_gaps_sheet(wb, matrix)
    _build_tool_coverage_sheet(wb, coverage_data, matrix)
    _build_methodology_sheet(wb, generated_at, model_used, coverage_data, excluded_tool_names or [])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
